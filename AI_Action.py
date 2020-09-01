# 20200901164001
from openpyxl import load_workbook
from openpyxl.styles import PatternFill  # 导入填充模块
from openpyxl import Workbook
import datetime
import time
import json
import requests
from pyquery import PyQuery as pq

# 域名与建站所有产品campaignId
campaignId_lists = []  # id总表
price_lists = []  # 价格总表
old_lists = []  # 原价总表

# 后端价格汇总
Price_new = []
Price_old = []
Price_campaignId = []

Project_name = 'AI产品'
Time1 = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')

Ym_api = 'https://cloud.baidu.com/api/yunying/discount/login/info/order'  # 接口地址
Page_Path = "https://cloud.baidu.com/campaign/Promotionai/index.html?unifrom=eventpage"  # 前端地址
page_headers = {
    'Cookie': 'AGL_USER_ID=5ed63b12-4440-4d47-b881-0f098f90410e; BIDUPSID=A9069B1FBC1A91044A9BEFCC8CFF8FE1; PSTM=1595216240; BAIDUID=965EAC8334F751D924041185AA64323A:FG=1; MCITY=-%3A; BDORZ=B490B5EBF6F3CD402E515D22BCDA1598; yjs_js_security_passport=e700866e4721bb2b3ec1695c399adf5930be8bcd_1597302603_js; BIDUPSID_BFESS=A9069B1FBC1A91044A9BEFCC8CFF8FE1; _ga=GA1.2.1775422955.1597372647; _gid=GA1.2.842133575.1597372647; BDUSS=0VNZm1JanVZfjg1SDh0Nm5JaC1ZdzJIUnlHOVExalpEb3Z5bExYVVdqWVBobDFmRUFBQUFBJCQAAAAAAAAAAAEAAACFA67ysOvU2LzT0rsAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA~5NV8P-TVfZ; BDUSS_BFESS=0VNZm1JanVZfjg1SDh0Nm5JaC1ZdzJIUnlHOVExalpEb3Z5bExYVVdqWVBobDFmRUFBQUFBJCQAAAAAAAAAAAEAAACFA67ysOvU2LzT0rsAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA~5NV8P-TVfZ; Hm_lvt_28a17f66627d87f1d046eae152a1c93d=1597198709,1597283859,1597372642,1597372691; UUAP_P_TOKEN=PT-500991099228983296-4YpivKwajF-uuap; SECURE_UUAP_P_TOKEN=PT-500991099228983296-4YpivKwajF-uuap; jsdk-user=v_dingguoqiang; BSG_B_TOKEN=3UaasTcl7sy9jCzBPHgOzkX6UJE7iVP9XKeDG8n0YnFgx/lTags+43ZWLjbMRwQ6gS6Qvrhs1t+WPya4p9rqJF/TK8dPECuzr9W54ca6xbc=; SECURE_BSG_B_TOKEN=3UaasTcl7sy9jCzBPHgOzkX6UJE7iVP9XKeDG8n0YnFgx/lTags+43ZWLjbMRwQ6gS6Qvrhs1t+WPya4p9rqJF/TK8dPECuzr9W54ca6xbc=; RT="z=1&dm=baidu.com&si=hkiudauq1p&ss=kdtnx7sd&sl=9&tt=axb&bcn=https%3A%2F%2Ffclog.baidu.com%2Flog%2Fweirwood%3Ftype%3Dperf&ld=do9a&cl=2ru5"; bce-userbind-source=PASSPORT%3BUUAP; bce-ctl-client-cookies="__cas__st__285,__cas__id__285,__cas__rn__,SIGNIN_UC,bce-device-cuid,bce-device-token,BAIDUID,ucbearer-clientid,ucbearer-devicecode,ucbearer-token,ucbearer-ucid"; bce-auth-type=UC; __cas__id__285=30736364; __cas__rn__=0; bce-sessionid=0028e62007484fe43f5a6b447d1c4ef7703; bce-user-info=2020-08-14T11:33:31Z|16f0e314d576f2a8e4bf4291adda89a4; bce-ctl-sessionmfa-cookie=bce-session; bce-session=db5adbbd3653492689b2bad486e13550743e6344eb9f40998fd9e19e26d35b04|0495112c3c478c5f86c9da6d4dd3c130; bce-login-type=UC; __cas__st__285=c838a4d3c17e3d4bc1025c773fb8cdf4485009ba450fe14bfc611a24070ee7a9750ead6fcc2ecfdb9a07dd65; bce-login-display-name=15011112222; _gat_gtag_UA_138572523_1=1; Hm_lpvt_28a17f66627d87f1d046eae152a1c93d=1597376027; BAIDU_CLOUD_TRACK_PATH=https%3A%2F%2Fcloudtest.baidu.com%2Fcampaign%2Forder.html%3F_%3D1597376011466%26unifrom%3Deventpage%23%2FaiDay%2FactivityPromotion~product%3Dnlp20200715%26config%3Dkill_nlp_540_10w_12%26qps%3D1'    # 'csrftoken': '2020-08-13T18:37:18Z|4ed1e831797bd2b9df7e60135c888f5e',
    # 'Content-Type': 'application/json'
}
api_headers = {
    'Cookie': 'AGL_USER_ID=5ed63b12-4440-4d47-b881-0f098f90410e; BIDUPSID=A9069B1FBC1A91044A9BEFCC8CFF8FE1; PSTM=1595216240; BAIDUID=965EAC8334F751D924041185AA64323A:FG=1; MCITY=-%3A; BDORZ=B490B5EBF6F3CD402E515D22BCDA1598; yjs_js_security_passport=e700866e4721bb2b3ec1695c399adf5930be8bcd_1597302603_js; BIDUPSID_BFESS=A9069B1FBC1A91044A9BEFCC8CFF8FE1; _ga=GA1.2.1775422955.1597372647; _gid=GA1.2.842133575.1597372647; BDUSS=0VNZm1JanVZfjg1SDh0Nm5JaC1ZdzJIUnlHOVExalpEb3Z5bExYVVdqWVBobDFmRUFBQUFBJCQAAAAAAAAAAAEAAACFA67ysOvU2LzT0rsAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA~5NV8P-TVfZ; BDUSS_BFESS=0VNZm1JanVZfjg1SDh0Nm5JaC1ZdzJIUnlHOVExalpEb3Z5bExYVVdqWVBobDFmRUFBQUFBJCQAAAAAAAAAAAEAAACFA67ysOvU2LzT0rsAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA~5NV8P-TVfZ; Hm_lvt_28a17f66627d87f1d046eae152a1c93d=1597198709,1597283859,1597372642,1597372691; UUAP_P_TOKEN=PT-500991099228983296-4YpivKwajF-uuap; SECURE_UUAP_P_TOKEN=PT-500991099228983296-4YpivKwajF-uuap; jsdk-user=v_dingguoqiang; BSG_B_TOKEN=3UaasTcl7sy9jCzBPHgOzkX6UJE7iVP9XKeDG8n0YnFgx/lTags+43ZWLjbMRwQ6gS6Qvrhs1t+WPya4p9rqJF/TK8dPECuzr9W54ca6xbc=; SECURE_BSG_B_TOKEN=3UaasTcl7sy9jCzBPHgOzkX6UJE7iVP9XKeDG8n0YnFgx/lTags+43ZWLjbMRwQ6gS6Qvrhs1t+WPya4p9rqJF/TK8dPECuzr9W54ca6xbc=; RT="z=1&dm=baidu.com&si=hkiudauq1p&ss=kdtnx7sd&sl=9&tt=axb&bcn=https%3A%2F%2Ffclog.baidu.com%2Flog%2Fweirwood%3Ftype%3Dperf&ld=do9a&cl=2ru5"; bce-userbind-source=PASSPORT%3BUUAP; bce-ctl-client-cookies="__cas__st__285,__cas__id__285,__cas__rn__,SIGNIN_UC,bce-device-cuid,bce-device-token,BAIDUID,ucbearer-clientid,ucbearer-devicecode,ucbearer-token,ucbearer-ucid"; bce-auth-type=UC; __cas__id__285=30736364; __cas__rn__=0; bce-sessionid=0028e62007484fe43f5a6b447d1c4ef7703; bce-user-info=2020-08-14T11:33:31Z|16f0e314d576f2a8e4bf4291adda89a4; bce-ctl-sessionmfa-cookie=bce-session; bce-session=db5adbbd3653492689b2bad486e13550743e6344eb9f40998fd9e19e26d35b04|0495112c3c478c5f86c9da6d4dd3c130; bce-login-type=UC; __cas__st__285=c838a4d3c17e3d4bc1025c773fb8cdf4485009ba450fe14bfc611a24070ee7a9750ead6fcc2ecfdb9a07dd65; bce-login-display-name=15011112222; _gat_gtag_UA_138572523_1=1; Hm_lpvt_28a17f66627d87f1d046eae152a1c93d=1597376027; BAIDU_CLOUD_TRACK_PATH=https%3A%2F%2Fcloudtest.baidu.com%2Fcampaign%2Forder.html%3F_%3D1597376011466%26unifrom%3Deventpage%23%2FaiDay%2FactivityPromotion~product%3Dnlp20200715%26config%3Dkill_nlp_540_10w_12%26qps%3D1',
    # 'csrftoken': '2020-08-13T18:37:18Z|4ed1e831797bd2b9df7e60135c888f5e',
    'Content-Type': 'application/json'
}

file_name = r'E:\BD_A\%s价格核对%s.xlsx' % (Project_name, Time1)
wb = Workbook()  # 实例对象需要wb.save(r'E:\BD_A\域名价格核对%s.xlsx'% Time1)来创建文件保存数据

# sheet重命名
sheet = wb['Sheet']
sheet.title = f'{Project_name}前后端价格对比'
Fille = PatternFill('solid', fgColor='FFBB02')  # solid 是填充实色的意思？
error_Fille = PatternFill('solid', fgColor='FF0000')  # darkVertical 是填充竖线  FF0000是红色
header_Fille = PatternFill('solid', fgColor='99FF00')  # darkVertical 是填充竖线  99FF00是黄色

# 先写头部
excel_header = ['id', '前端campaignId', '前端价格', '后端价格', '对比结果', '前端原价', '后端原价', '后端campaignId']  # 首行头部信息作为变量可更改
for i in range(len(excel_header)):
    sheet.cell(row=1, column=i+1).value = excel_header[i]
    sheet.cell(row=1, column=i+1).fill = header_Fille  # 首行颜色
    sheet.freeze_panes = wb.active.cell(row=2, column=1)  # 首行冻结

    # 单元格宽度设置
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 45
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 20  # 对比结果
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 45

data = {}
Page_M = requests.request("get", url=Page_Path, data=data, headers=page_headers)
Page_M.encoding = 'UTF-8'
PQ_Page = pq(Page_M.text)

print('前端信息')
print('='*100)
# 新用户限量抢购
New_ids = []
request_New_id = []  # 需要拼接的字段为这个
request_New_name = []

for i in PQ_Page('.module.module-0 .ai-common-mod1.product'):
    New_id = pq(i).find('.btn').attr('data-id')  # id
    New_ids.append(New_id)
    New_data = json.loads(pq(i).find('.value .v-select').attr('data-datasource'))
    [campaignId_lists.append(New_id + '_' + h) for h in [New_data[j]['value'] for j in range(len(New_data))]]
    [request_New_name.append(h) for h in [New_data[j]['name'] for j in range(len(New_data))]]
    [old_lists.append(h) for h in [New_data[j]['old'] for j in range(len(New_data))]]
    [price_lists.append(h) for h in [New_data[j]['price'] for j in range(len(New_data))]]


# 精选特惠
Preferential_ids = []
for i in PQ_Page('.module.module-1 .ai-common-mod2.product'):
    Preferential_id = pq(i).find('.btn').attr('data-id')
    try:
        Preferential_data = json.loads(pq(i).find('.value .v-select').attr('data-datasource'))
        [campaignId_lists.append(Preferential_id + '_' + h) for h in [Preferential_data[j]['value'] for j in range(len(Preferential_data))]]
        [old_lists.append(h) for h in [Preferential_data[j]['old'] for j in range(len(Preferential_data))]]
        [price_lists.append(h) for h in [Preferential_data[j]['price'] for j in range(len(Preferential_data))]]
    except TypeError as e:
        Preferential_id = pq(i).find('.btn').attr('data-id')
        Preferential_old = pq(i).find('.btn').attr('data-old')
        Preferential_price = pq(i).find('.btn').attr('data-price')
        if '230' in pq(i).find('.select-item').attr('data-id'):
            pass
            # campaignId_lists.append(Preferential_id)
            # old_lists.append(Preferential_old)
            # price_lists.append(Preferential_price)
        else:
            for k in pq(i).find('.select-item'):
                pass
                # campaignId_lists.append(pq(i).find('.btn').attr('data-id')+'_'+pq(k).attr('data-id'))
                # old_lists.append(pq(k).attr('data-price'))
                # price_lists.append(pq(k).attr('data-old'))

x = 0
for i in range(len(campaignId_lists)):
    x += 1
    sheet.cell(row=i + 2, column=1).value = x
    sheet.cell(row=i + 2, column=2).value = campaignId_lists[i]

for i in range(len(price_lists)):
    sheet.cell(row=i + 2, column=3).value = str(price_lists[i])
    sheet.cell(row=i + 2, column=3).fill = Fille  # 将当前行的每一个表格填充颜色

for i in range(len(old_lists)):
    sheet.cell(row=i + 2, column=6).value = str(old_lists[i])
    sheet.cell(row=i + 2, column=6).fill = Fille  # 将当前行的每一个表格填充颜色

print('前端信息导入完毕！')
print('=' * 100)


data_count = 0
for i in campaignId_lists:
    data_count += 1
    data = "{\"campaignId\":\"%s\"}" % i  # 阔以请求到数据
    print(data_count,data)

    # data = {"campaignId":"20200607_discount_kill_aipage_site_miniprogram_1_12m"}  # 报500错误
    count = 0

    while True:
        try:
            # 请求的时候参数必须为url,data,headers。如果headers内需要其他参数的话，必须要用headers=headers，直接写cookies=XX会报错
            result = requests.post(url=Ym_api, data=data, headers=api_headers).json()  # verify=False

        except requests.exceptions.ProxyError as e:
            print(e)
            print('远程访问被拒绝，正在重连...')
        except requests.exceptions.ConnectionError as r:
            print(r, f'\n已经{count}分钟')
            time.sleep(60)
            count += 1
        else:
            count += 1
            try :
                if result['result']['price']:
                    Price_new.append(str(result['result']['price']['campaignPrice']).split('.')[0])
                    Price_old.append(str(result['result']['price']['originalPrice']).split('.')[0])
                    Price_campaignId.append(result['result']['privilegeId'])
            except KeyError as e:
                Price_new.append('本产品查询报500')
                if result['error']:
                    Price_old.append(result['error'])
                else:
                    Price_old.append('缺少报错信息')
                Price_campaignId.append(data)
                print(result,'价格没有查出来！', data)
                pass
            break

    # 调试第二条数据跳出
    if data_count == 2:
        break
for i in range(len(Price_new)):
    sheet.cell(row=i + 2, column=4).value = str(Price_new[i])
    sheet.cell(row=i + 2, column=4).fill = Fille  # 将当前行的每一个表格填充颜色

for i in range(len(Price_old)):
    sheet.cell(row=i + 2, column=7).value = str(Price_old[i])
    sheet.cell(row=i + 2, column=7).fill = Fille  # 将当前行的每一个表格填充颜色

for i in range(len(Price_campaignId)):
    sheet.cell(row=i + 2, column=8).value = Price_campaignId[i]
print('后端信息导入完毕')
wb.save(file_name)

# =========读取数据=========
print('开始对比前后端数据...')


def read_excel(file_path, sheet_name):
    # 读取Excel表格表面数据，不读取公式，如需获取表格内公式，将data_only改为false(本代码不推荐，因为改了之后不会显示Excel数据)
    file_sheet = load_workbook(file_path, data_only=True,)[sheet_name]

    # 获取头部
    sheet_header = []
    for Header_data in range(file_sheet.max_column):
        sheet_header.append(file_sheet.cell(row=1, column=Header_data+1).value)

    # 获取数据
    data_list = []
    for data_data1 in range(file_sheet.max_row-1):
        data_dict = {}
        for data_data2 in range(len(sheet_header)):
            data_dict[sheet_header[data_data2]] = file_sheet.cell(row=data_data1+2, column=data_data2+1).value
        data_list.append(data_dict)
    return data_list  # 返回数据

# =======读取数据并且在指定列追加数据=======
file_Data = read_excel(file_name, sheet.title)
wb1 = load_workbook(file_name)
ws1 = wb1[sheet.title]

# =========数据对比=========
m = 0
for i in file_Data:
    if (i['前端价格'] == i['后端价格']) and i['前端原价'] == i['后端原价']:
        m += 1
        ws1.cell(row=m + 1, column=5).value = "前后端价格一致"  # excel追加数据

    elif (i['前端价格'] is None) or (i['前端原价'] is None):
        print(i)
        m += 1
        ws1.cell(row=m + 1, column=5).value = "没有查询到前端数据"  # excel追加数据
        ws1.cell(row=m + 1, column=5).fill = error_Fille  # 将当前行的每一个表格填充颜色
    elif (i['后端价格'] is None) or (i['后端原价'] is None):
        print(i)
        m += 1
        ws1.cell(row=m + 1, column=5).value = "没有查询到后端数据"  # excel追加数据
        ws1.cell(row=m + 1, column=5).fill = error_Fille  # 将当前行的每一个表格填充颜色

    elif i['前端价格'] != i['后端价格']:
        print(i)
        m += 1
        ws1.cell(row=m + 1, column=5).value = "售卖价格不一致"  # excel追加数据
        ws1.cell(row=m + 1, column=5).fill = error_Fille  # 将当前行的每一个表格填充颜色

    elif i['前端原价'] != i['后端原价']:
        print(i)
        m += 1
        ws1.cell(row=m + 1, column=5).value = "原价不一致"  # excel追加数据
        ws1.cell(row=m + 1, column=5).fill = error_Fille  # 将当前行的每一个表格填充颜色

wb1.save(file_name)
